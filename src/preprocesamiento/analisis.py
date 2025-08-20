import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import pandas as pd

def generar_reporte_excel(df_limpio, conteo, ruta="reporte_spotify.xlsx", imagen="top_artistas.png"):
    # Guardar datos en Excel
    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        df_limpio.to_excel(writer, sheet_name="Datos Limpios", index=False)
        conteo.to_excel(writer, sheet_name="Top Artistas")

    # Insertar gráfico
    wb = load_workbook(ruta)
    ws = wb.create_sheet("Graficos")
    img = Image(imagen)
    ws.add_image(img, "A1")
    wb.save(ruta)
    
    print(f"✅ Reporte generado: {ruta}")


def crear_carpeta_reportes(subcarpeta=None):
    carpeta_base = os.path.join("data", "raw", "reportes")
    if subcarpeta:
        carpeta_base = os.path.join(carpeta_base, subcarpeta)
    os.makedirs(carpeta_base, exist_ok=True)
    return carpeta_base

def top_artistas(df, carpeta_reportes):
    top_artistas = df['artist_name'].value_counts().head(5)
    print("\n🎤 Top 5 artistas más escuchados:")
    print(top_artistas)
    top_artistas.plot(kind='barh', title="Top 5 Artistas más escuchados", color='skyblue')
    plt.xlabel("Reproducciones")
    plt.ylabel("Artista")
    plt.tight_layout()
    plt.savefig(os.path.join(carpeta_reportes, "top_artistas.png"))
    plt.close()

def top_canciones(df, carpeta_reportes):
    top_tracks = df.groupby('track_name')['ms_played'].sum().sort_values(ascending=False).head(5)
    print("\n🎵 Top 5 canciones por tiempo total reproducido:")
    print(top_tracks)
    (top_tracks / 60000).plot(kind='barh', title="Top 5 Canciones por tiempo reproducido (min)", color='orange')
    plt.xlabel("Minutos reproducidos")
    plt.ylabel("Canción")
    plt.tight_layout()
    plt.savefig(os.path.join(carpeta_reportes, "top_canciones.png"))
    plt.close()

def canciones_saltadas(df, carpeta_reportes):
    skipped_counts = df['skipped'].value_counts()
    print("\n⏭️ Cantidad de canciones saltadas:")
    print(skipped_counts)
    skipped_counts.plot(kind='pie', autopct='%1.1f%%', title="Porcentaje de canciones saltadas", colors=["#66c2a5", "#fc8d62"])
    plt.ylabel("")
    plt.tight_layout()
    plt.savefig(os.path.join(carpeta_reportes, "canciones_saltadas.png"))
    plt.close()

def resumen(df):
    carpeta = crear_carpeta_reportes()
    top_artistas(df, carpeta)
    top_canciones(df, carpeta)
    canciones_saltadas(df, carpeta)
    print(f"\n📂 Gráficos guardados en la carpeta '{carpeta}'")
